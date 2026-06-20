"""
Report service — PDF and Excel generation for all modules
"""
import io
from datetime import datetime, timezone
from typing import Optional
from sqlalchemy.orm import Session

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models.sales import SalesOrder, SalesOrderItem
from app.models.purchase import PurchaseOrder, PurchaseOrderItem
from app.models.product import Product
from app.models.manufacturing import ManufacturingOrder


def _header_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])


def generate_sales_report_pdf(
    db: Session,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch,
                             topMargin=0.75*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Mini ERP — Sales Report", styles['h1']))
    elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))

    q = db.query(SalesOrder).filter(SalesOrder.deleted_at.is_(None))
    if start_date:
        q = q.filter(SalesOrder.created_at >= start_date)
    if end_date:
        q = q.filter(SalesOrder.created_at <= end_date)
    orders = q.order_by(SalesOrder.created_at.desc()).all()

    data = [["Order #", "Customer", "Status", "Subtotal", "Tax", "Total", "Date"]]
    for o in orders:
        data.append([
            o.order_number,
            o.customer.name if o.customer else "N/A",
            o.status.upper(),
            f"${o.subtotal:,.2f}",
            f"${o.tax_amount:,.2f}",
            f"${o.total_amount:,.2f}",
            o.created_at.strftime("%Y-%m-%d") if o.created_at else "",
        ])

    table = Table(data, colWidths=[1.2*inch, 1.8*inch, 1*inch, 1*inch, 0.8*inch, 1*inch, 1*inch])
    table.setStyle(_header_style())
    elements.append(table)

    total_rev = sum(o.total_amount for o in orders)
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph(f"<b>Total Revenue: ${total_rev:,.2f}</b>", styles['Normal']))

    doc.build(elements)
    return buf.getvalue()


def generate_sales_report_excel(
    db: Session,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Orders"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    headers = ["Order #", "Customer", "Status", "Subtotal", "Tax", "Total", "Date"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    q = db.query(SalesOrder).filter(SalesOrder.deleted_at.is_(None))
    if start_date:
        q = q.filter(SalesOrder.created_at >= start_date)
    if end_date:
        q = q.filter(SalesOrder.created_at <= end_date)

    for row, o in enumerate(q.order_by(SalesOrder.created_at.desc()).all(), 2):
        ws.cell(row=row, column=1, value=o.order_number)
        ws.cell(row=row, column=2, value=o.customer.name if o.customer else "")
        ws.cell(row=row, column=3, value=o.status.upper())
        ws.cell(row=row, column=4, value=round(o.subtotal, 2))
        ws.cell(row=row, column=5, value=round(o.tax_amount, 2))
        ws.cell(row=row, column=6, value=round(o.total_amount, 2))
        ws.cell(row=row, column=7, value=o.created_at.strftime("%Y-%m-%d") if o.created_at else "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_inventory_report_pdf(db: Session) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch,
                             topMargin=0.75*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Mini ERP — Inventory Report", styles['h1']),
        Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']),
        Spacer(1, 0.2*inch),
    ]

    products = db.query(Product).filter(Product.deleted_at.is_(None), Product.is_active == "true").all()
    data = [["SKU", "Name", "Stock", "Reserved", "Free", "Cost", "Value", "Reorder"]]
    total_value = 0.0
    for p in products:
        val = p.stock_qty * p.cost_price
        total_value += val
        free = max(0, p.stock_qty - p.reserved_qty)
        data.append([
            p.sku, p.name[:25],
            f"{p.stock_qty:.0f}", f"{p.reserved_qty:.0f}", f"{free:.0f}",
            f"${p.cost_price:.2f}", f"${val:,.2f}", f"{p.reorder_level:.0f}",
        ])

    table = Table(data, colWidths=[0.9*inch]*8)
    table.setStyle(_header_style())
    elements.append(table)
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph(f"<b>Total Inventory Value: ${total_value:,.2f}</b>", styles['Normal']))
    doc.build(elements)
    return buf.getvalue()


def generate_purchase_report_excel(db: Session) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    headers = ["Order #", "Vendor", "Status", "Total", "Expected", "Received"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    orders = db.query(PurchaseOrder).filter(PurchaseOrder.deleted_at.is_(None)).order_by(PurchaseOrder.created_at.desc()).all()
    for row, o in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=o.order_number)
        ws.cell(row=row, column=2, value=o.vendor.name if o.vendor else "")
        ws.cell(row=row, column=3, value=o.status.upper())
        ws.cell(row=row, column=4, value=round(o.total_amount, 2))
        ws.cell(row=row, column=5, value=o.expected_date.strftime("%Y-%m-%d") if o.expected_date else "")
        ws.cell(row=row, column=6, value=o.received_date.strftime("%Y-%m-%d") if o.received_date else "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_manufacturing_report_excel(db: Session) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manufacturing Orders"

    headers = ["MO #", "Product", "Planned Qty", "Produced Qty", "Status", "Start", "End"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    orders = db.query(ManufacturingOrder).filter(ManufacturingOrder.deleted_at.is_(None)).order_by(ManufacturingOrder.created_at.desc()).all()
    for row, o in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=o.mo_number)
        ws.cell(row=row, column=2, value=o.product.name if o.product else "")
        ws.cell(row=row, column=3, value=o.planned_qty)
        ws.cell(row=row, column=4, value=o.produced_qty)
        ws.cell(row=row, column=5, value=o.status.upper())
        ws.cell(row=row, column=6, value=o.actual_start.strftime("%Y-%m-%d") if o.actual_start else "")
        ws.cell(row=row, column=7, value=o.actual_end.strftime("%Y-%m-%d") if o.actual_end else "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
